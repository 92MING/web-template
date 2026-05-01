

import xlrd
import tempfile

from io import BytesIO
from pathlib import Path
from typing import Any, ClassVar, Sequence

from numbers_parser import Document as NumbersDocument
from odf import table as odf_table
from odf import teletype
from odf.opendocument import load as odf_load
from openpyxl import load_workbook

from ._structured import TableSheetLikedDocument
from .base import LLMDocumentPart


class Excel(TableSheetLikedDocument):
    '''表格文档模型。

    NOTE: 目前支援 `.xls` / `.xlsx` / `.xlsm` / `.ods` / `.numbers`。不在此范围内的格式在导入时会直接报错。
    '''

    Abstract: ClassVar[bool] = False
    Type: ClassVar[str] = 'excel'
    TypeNames: ClassVar[tuple[str, ...]] = ('spreadsheet', 'sheet', 'xls', 'xlsx', 'xlsm', 'ods', 'numbers')
    Suffixes: ClassVar[tuple[str, ...]] = ('.xls', '.xlsx', '.xlsm', '.xltx', '.xltm', '.ods', '.numbers')
    MimeTypes: ClassVar[tuple[str, ...]] = ('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)

    def _extract_xlsx_sheets(self) -> list[tuple[str, list[list[str]]]]:
        workbook = load_workbook(BytesIO(self.to_bytes()), read_only=True, data_only=False)
        try:
            return [
                (
                    sheet.title,
                    [
                        [str(cell) if cell is not None else '' for cell in row]
                        for row in sheet.iter_rows(values_only=True)
                    ],
                )
                for sheet in workbook.worksheets
            ]
        finally:
            workbook.close()

    def _extract_xls_sheets(self) -> list[tuple[str, list[list[str]]]]:
        workbook = xlrd.open_workbook(file_contents=self.to_bytes())
        sheets: list[tuple[str, list[list[str]]]] = []
        for sheet in workbook.sheets():
            rows = [[str(cell) if cell is not None else '' for cell in sheet.row_values(row_idx)] for row_idx in range(sheet.nrows)]
            sheets.append((sheet.name, rows))
        return sheets

    def _extract_ods_sheets(self) -> list[tuple[str, list[list[str]]]]:
        data = self.to_bytes()
        with tempfile.NamedTemporaryFile(delete=False, suffix='.ods') as tmp_file:
            tmp_file.write(data)
            tmp_path = Path(tmp_file.name)

        try:
            ods_doc = odf_load(str(tmp_path))
        finally:
            tmp_path.unlink(missing_ok=True)

        sheet_texts: list[tuple[str, list[list[str]]]] = []
        spreadsheet_root = getattr(ods_doc, 'spreadsheet', None)
        if spreadsheet_root is None:
            return []

        for table in spreadsheet_root.getElementsByType(odf_table.Table):
            sheet_name = str(table.getAttribute('name') or 'Sheet').strip() or 'Sheet'
            rows: list[list[str]] = []
            for row in table.getElementsByType(odf_table.TableRow):
                row_values: list[str] = []
                for cell in row.getElementsByType(odf_table.TableCell):
                    repeat_raw = cell.getAttribute('numbercolumnsrepeated') or '1'
                    try:
                        repeat = max(1, int(str(repeat_raw)))
                    except Exception:
                        repeat = 1
                    cell_text = str(teletype.extractText(cell) or '').strip()
                    row_values.extend([cell_text] * repeat)
                rows.append(row_values)
            sheet_texts.append((sheet_name, rows))
        return sheet_texts

    def _extract_numbers_sheets(self) -> list[tuple[str, list[list[str]]]]:
        data = self.to_bytes()
        with tempfile.NamedTemporaryFile(delete=False, suffix='.numbers') as tmp_file:
            tmp_file.write(data)
            tmp_path = Path(tmp_file.name)

        try:
            doc = NumbersDocument(str(tmp_path))
            sheets: list[tuple[str, list[list[str]]]] = []
            for sheet in doc.sheets:
                sheet_rows: list[list[str]] = []
                for table in sheet.tables:
                    for row in table.iter_rows():
                        row_values: list[str] = []
                        for cell in row:
                            formatted = getattr(cell, 'formatted_value', None)
                            value = getattr(cell, 'value', None)
                            cell_text = formatted if formatted not in (None, '') else value
                            row_values.append('' if cell_text is None else str(cell_text))
                        sheet_rows.append(row_values)
                sheets.append((sheet.name or 'Sheet', sheet_rows))
            return sheets
        finally:
            tmp_path.unlink(missing_ok=True)

    def to_sheets(self) -> list[tuple[str, list[list[str]]]]:
        suffix = self.suffix
        zip_names = {name.lower() for name in self._zip_file_names()}
        zip_mimetype = self._zip_mimetype()
        try:
            if suffix in {'.xlsx', '.xlsm', '.xltx', '.xltm'} or 'xl/workbook.xml' in zip_names:
                return self._extract_xlsx_sheets()
            elif suffix == '.xls':
                return self._extract_xls_sheets()
            elif suffix == '.ods' or 'opendocument.spreadsheet' in zip_mimetype:
                return self._extract_ods_sheets()
            elif suffix == '.numbers' or any(name.startswith('index/') or name.startswith('metadata/') for name in zip_names):
                return self._extract_numbers_sheets()
            else:
                raise ValueError(f'Unsupported spreadsheet format: {suffix or "unknown"}.')
        except Exception as exc:
            raise ValueError(f'Failed to parse spreadsheet: {type(exc).__name__}: {exc}') from exc

    async def to_llm(self, **kwargs: Any) -> Sequence[LLMDocumentPart]:
        final_text = self.sheets_to_text(delimiter='\t')
        if not final_text:
            final_text = '#Sheet1\n'
        return [final_text]


__all__ = ['Excel']
